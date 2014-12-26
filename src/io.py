import csv
import sys
from openpyxl import load_workbook, Workbook


def save_to_csv(data_set, out=sys.stdout):
    writer = csv.writer(out, lineterminator='\n')
    for element in data_set:
        class_number = element[0]
        attributes = element[1]
        writer.writerow([class_number] + attributes)


def load_from_csv(csv_file, attr_type=float):
    reader = csv.reader(csv_file)
    data_set = []
    for row in reader:
        class_number = int(row[0])
        attributes = [attr_type(attribute) for attribute in row[1:]]
        data_set.append([class_number, attributes])
    return data_set


def load_from_csv_file(file, attr_type=float):
    return load_from_csv(open(file, "r"), attr_type)


def load_from_xlsx_file(file, attr_type=float):
    wb = load_workbook(filename=file, use_iterators=True)
    ws = wb.active
    data_set = []
    for i, row in enumerate(ws.iter_rows(), start=1):
        row = [x.value for x in row]
        class_number = int(row[0])
        attributes = [attr_type(attribute) for attribute in row[1:]]
        data_set.append([class_number, attributes])
        if not i % 1000:
            print("{} rows loaded...".format(i))
    return data_set


def load_file(file, attr_type=float):
    if file.endswith(".xlsx"):
        return load_from_xlsx_file(file, attr_type)
    elif file.endswith(".csv"):
        return load_from_csv_file(file, attr_type)
    else:
        raise ValueError


def save_classes_to_xlsx_file(classes, file):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    for c in classes:
        ws.append([c])
    wb.save(file)


def save_classes_file(classes, file):
    if file.endswith(".xlsx"):
        return save_classes_to_xlsx_file(classes, file)
    else:
        raise ValueError
